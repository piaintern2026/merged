"""
routes/intern.py
-----------------
Full CRUD for Intern registration, including creating their linked User
login account, profile picture upload and a detail "view" page.
"""

import csv
import io
from datetime import datetime, date

from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, send_file, session
from flask_login import login_required
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError

from extensions import db
from models import (
    Intern,
    Department,
    SubDepartment,
    User,
    Attendance,
    Project,
    ProjectSubmission,
    FinalReport,
    Feedback,
    PMEvaluation,
    ProjectManager,
)
from utils import (
    roles_required,
    save_profile_picture,
    delete_profile_picture,
    log_action,
    notify_user,
    PIA_CITIES,
    CNIC_RE,
    EMAIL_RE,
    PHONE_RE,
    normalize_cnic,
    normalize_phone,
    hr_city_scope,
    DEFAULT_USER_PASSWORD,
)
from services.email_service import send_welcome_email, send_hr_pm_notification_email, get_hr_recipients

intern_bp = Blueprint("intern", __name__, url_prefix="/interns")


def _parse_date(value: str):
    """Parse an HTML date input (YYYY-MM-DD) into a date object."""
    return datetime.strptime(value, "%Y-%m-%d").date()


def _get_or_create_department(placement_name: str):
    """Resolve a Department row for the given free-text Placement value
    (case-insensitive match on name), creating one if none exists yet.

    The Excel template's "Placement" column doesn't require the admin
    to pre-create a matching Department, but the app's Project
    assignment, Rotation Management, and Attendance/Leave filtering
    all key off Intern.department_id -- so every intern is still
    attached to a real Department row, just derived automatically
    instead of picked from a required dropdown."""
    name = (placement_name or "").strip()
    if not name:
        return None
    department = Department.query.filter(db.func.lower(Department.name) == name.lower()).first()
    if department:
        return department
    department = Department(name=name, is_active=True)
    db.session.add(department)
    db.session.flush()
    return department


def _create_intern_account(
    *,
    full_name,
    cnic,
    university,
    qualification,
    major,
    placement,
    station,
    phone,
    email,
    username,
    password,
    start_date,
    end_date,
    documents_status="Pending",
    certificate_status="Pending",
    semester=None,
    profile_picture_filename=None,
    force_password_reset=False,
    department=None,
    sub_department_id=None,
):
    """Create the linked User + Intern rows.

    This is the single source of truth for "registering an intern" --
    used by both the manual Add Intern form (which now picks a real
    Department from a dropdown, passed in directly as `department`)
    and the Bulk Import Interns (Excel/CSV) feature, which still only
    has a free-text "Placement" column and so falls back to the
    get-or-create lookup -- so the two paths can never drift apart
    (same fields, same password hashing, same linking).
    """
    user = User(
        email=email,
        username=username,
        role="Intern",
        profile_picture=profile_picture_filename,
        force_password_reset=force_password_reset,
    )
    user.set_password(password)
    db.session.add(user)
    db.session.flush()

    if department is None:
        department = _get_or_create_department(placement)

    intern = Intern(
        user_id=user.id,
        full_name=full_name,
        cnic=cnic,
        university=university,
        qualification=qualification,
        major=major or None,
        semester=semester or None,
        department_id=department.id if department else None,
        sub_department_id=sub_department_id,
        placement=placement or None,
        station=station,
        phone=phone,
        internship_start_date=start_date,
        internship_end_date=end_date,
        documents_status=documents_status or "Pending",
        certificate_status=certificate_status or "Pending",
    )
    db.session.add(intern)
    db.session.flush()
    return user, intern


@intern_bp.route("/")
@login_required
@roles_required("Station HR", "Admin")
def list_interns():
    """Show all registered interns, with client-side search/filter by
    name/CNIC, city, university, department, assigned Project Manager
    and internship status, plus a server-side year filter (by
    internship start date)."""
    from sqlalchemy import extract

    year = request.args.get("year", type=int)
    is_city_scoped, hr_city = hr_city_scope()

    query = Intern.query
    if is_city_scoped:
        query = query.filter(Intern.station == hr_city)
    if year:
        query = query.filter(extract("year", Intern.internship_start_date) == year)

    interns = query.order_by(Intern.created_at.desc()).all()

    available_years = sorted(
        {
            y[0]
            for y in db.session.query(extract("year", Intern.internship_start_date)).distinct().all()
            if y[0] is not None
        },
        reverse=True,
    )

    universities = sorted({i.university for i in interns if i.university}, key=str.lower)
    qualifications = sorted({i.qualification for i in interns if i.qualification}, key=str.lower)
    majors = sorted({i.major for i in interns if i.major}, key=str.lower)
    departments = Department.query.order_by(db.func.lower(Department.name)).all()
    managers = ProjectManager.query.order_by(db.func.lower(ProjectManager.full_name)).all()
    if is_city_scoped:
        departments = [d for d in departments if d.city == hr_city]
        managers = [m for m in managers if m.city == hr_city]
    return render_template(
        "interns/list.html",
        interns=interns,
        stations=[hr_city] if is_city_scoped and hr_city else PIA_CITIES,
        universities=universities,
        qualifications=qualifications,
        majors=majors,
        departments=departments,
        managers=managers,
        statuses=Intern.INTERNSHIP_STATUSES,
        documents_statuses=Intern.DOCUMENTS_STATUSES,
        certificate_statuses=Intern.CERTIFICATE_STATUSES,
        available_years=available_years,
        selected_year=year,
    )


@intern_bp.route("/view/<int:intern_id>")
@login_required
@roles_required("Station HR", "Admin")
def view_intern(intern_id):
    """Show full details of a single intern, including their Module 3
    portal activity: assigned project, submissions, work logs, final
    report, feedback and attendance summary. A Station HR may only view
    interns whose station/city matches their own assigned city."""
    intern = Intern.query.get_or_404(intern_id)

    is_city_scoped, hr_city = hr_city_scope()
    if is_city_scoped and intern.station != hr_city:
        flash("You do not have permission to view interns outside your assigned city.", "danger")
        return redirect(url_for("intern.list_interns"))

    assigned_projects = (
        Project.query.filter(Project.interns.any(id=intern.id))
        .order_by(Project.created_at.desc())
        .all()
    )

    attendance_records = Attendance.query.filter_by(intern_id=intern.id).all()
    total_attendance = len(attendance_records)
    present_count = sum(1 for r in attendance_records if r.status in Attendance.ATTENDED_STATUSES)
    attendance_percentage = (
        round((present_count / total_attendance) * 100, 1) if total_attendance else 0
    )

    submissions = (
        ProjectSubmission.query.filter_by(intern_id=intern.id)
        .order_by(ProjectSubmission.submitted_at.desc())
        .all()
    )
    final_report = FinalReport.query.filter_by(intern_id=intern.id).first()
    feedback = Feedback.query.filter_by(intern_id=intern.id).first()
    pm_evaluations = (
        PMEvaluation.query.filter_by(intern_id=intern.id)
        .order_by(PMEvaluation.evaluation_date.desc())
        .all()
    )

    return render_template(
        "interns/view.html",
        intern=intern,
        assigned_projects=assigned_projects,
        attendance_percentage=attendance_percentage,
        total_attendance=total_attendance,
        submissions=submissions,
        final_report=final_report,
        feedback=feedback,
        pm_evaluations=pm_evaluations,
    )


@intern_bp.route("/add", methods=["GET", "POST"])
@login_required
@roles_required("Admin")
def add_intern():
    """Register a new intern (creates User + Intern rows)."""

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        cnic = request.form.get("cnic", "").strip()
        station = request.form.get("station", "").strip()
        phone = request.form.get("phone", "").strip()
        email = request.form.get("email", "").strip().lower()
        qualification = request.form.get("qualification", "").strip()
        major = request.form.get("major", "").strip()
        university = request.form.get("university", "").strip()
        department_id_raw = request.form.get("department_id", "").strip()
        sub_department_id_raw = request.form.get("sub_department_id", "").strip()
        documents_status = request.form.get("documents_status", "").strip()
        certificate_status = request.form.get("certificate_status", "").strip()
        username = request.form.get("username", "").strip() or email  # fallback: email becomes username
        password = request.form.get("password", "")
        start_date_raw = request.form.get("internship_start_date", "")
        end_date_raw = request.form.get("internship_end_date", "")
        photo = request.files.get("profile_picture")

        # ---- Validation ----
        errors = []
        required_fields = {
            "Full Name": full_name,
            "CNIC": cnic,
            "Station": station,
            "Cell No": phone,
            "Email": email,
            "Qualification": qualification,
            "Major": major,
            "University": university,
            "Department": department_id_raw,
            "Username": username,
            "Internship start date": start_date_raw,
            "Internship end date": end_date_raw,
        }
        for label, value in required_fields.items():
            if not value:
                errors.append(f"{label} is required.")

        department = None
        if department_id_raw:
            try:
                department = Department.query.get(int(department_id_raw))
            except ValueError:
                department = None
            if department is None:
                errors.append("Please select a valid department from the list.")

        sub_department = None
        if sub_department_id_raw:
            try:
                sub_department = SubDepartment.query.get(int(sub_department_id_raw))
            except ValueError:
                sub_department = None
            if sub_department is None or (
                department and sub_department.department_id != department.id
            ):
                errors.append("Please select a valid sub department for the chosen department.")
                sub_department = None

        if station and station not in PIA_CITIES:
            errors.append("Please select a valid station from the list.")

        if cnic and not CNIC_RE.match(cnic.replace(" ", "")):
            errors.append("CNIC must be in the format 42101-1234567-1 (13 digits).")
        if email and not EMAIL_RE.match(email):
            errors.append("Please enter a valid email address.")
        if phone and not PHONE_RE.match(normalize_phone(phone)):
            errors.append("Please enter a valid Pakistani cell number (e.g. 03001234567).")

        if documents_status and documents_status not in Intern.DOCUMENTS_STATUSES:
            errors.append("Please select a valid Documents Status.")
        if certificate_status and certificate_status not in Intern.CERTIFICATE_STATUSES:
            errors.append("Please select a valid Certificate Status.")

        # Optional: if the admin leaves this blank, the account gets the
        # standard default password (utils.DEFAULT_USER_PASSWORD) instead
        # of requiring one to be typed. A typed password still has to
        # meet the normal minimum length.
        if password and len(password) < 8:
            errors.append("Password must be at least 8 characters long.")

        cnic_normalized = normalize_cnic(cnic) if cnic else cnic

        if User.query.filter_by(email=email).first():
            errors.append("A user with this email already exists.")
        if User.query.filter_by(username=username).first():
            errors.append("This username is already taken.")
        if cnic_normalized and Intern.query.filter_by(cnic=cnic_normalized).first():
            errors.append("This CNIC is already registered.")

        start_date = end_date = None
        if start_date_raw and end_date_raw and not errors:
            try:
                start_date = _parse_date(start_date_raw)
                end_date = _parse_date(end_date_raw)
                if start_date > end_date:
                    errors.append("Internship start date must not be after the end date.")
            except ValueError:
                errors.append("Invalid date format provided.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template(
                "interns/form.html", intern=None, stations=PIA_CITIES, form=request.form,
                documents_statuses=Intern.DOCUMENTS_STATUSES, certificate_statuses=Intern.CERTIFICATE_STATUSES,
                departments=Department.query.filter_by(is_active=True).order_by(db.func.lower(Department.name)).all(),
            )

        try:
            picture_filename = save_profile_picture(photo)
            placement = department.name if department else None

            user, intern = _create_intern_account(
                full_name=full_name,
                cnic=cnic_normalized,
                university=university,
                qualification=qualification,
                major=major,
                placement=placement,
                department=department,
                sub_department_id=sub_department.id if sub_department else None,
                station=station,
                phone=normalize_phone(phone),
                email=email,
                username=username,
                password=password or DEFAULT_USER_PASSWORD,
                start_date=start_date,
                end_date=end_date,
                documents_status=documents_status,
                certificate_status=certificate_status,
                profile_picture_filename=picture_filename,
            )
            log_action(
                action="CREATE",
                description=f"Registered intern '{full_name}' (CNIC {cnic_normalized}).",
                target_type="Intern",
                target_id=intern.id,
            )
            db.session.commit()

            # In-app notifications for HR/Admin so a new registration
            # doesn't go unnoticed even if email delivery fails.
            for staff in User.query.filter(
                User.role.in_(["Station HR", "Admin"]), User.is_active_account.is_(True)
            ).all():
                notify_user(
                    staff.id,
                    f"New intern '{full_name}' has been registered.",
                    icon="bi-person-plus",
                    notification_type="General",
                )
            db.session.commit()

            # Email notifications - fired only after the commit above
            # succeeded. Failures are logged internally and never
            # interrupt this request (see services/email_service.py).
            send_welcome_email(user=user, intern=intern, raw_password=password)
            send_hr_pm_notification_email(
                recipients=get_hr_recipients(),
                recipient_name="HR Team",
                event_title="New Intern Registered",
                event_message=f"A new intern, {full_name}, has been registered in the system.",
                details=[
                    ("Name", full_name),
                    ("Department", placement or "N/A"),
                    ("CNIC", cnic_normalized),
                    ("Internship Period", f"{start_date} to {end_date}"),
                ],
            )

            flash(f"Intern '{full_name}' registered successfully.", "success")
            return redirect(url_for("intern.list_interns"))

        except ValueError as ve:
            db.session.rollback()
            flash(str(ve), "danger")
        except IntegrityError:
            db.session.rollback()
            flash("Could not register intern due to a database error.", "danger")
        except Exception:
            db.session.rollback()
            current_app.logger.exception("Failed to register new intern.")
            flash("Could not register intern due to a system error. Please try again.", "danger")

    return render_template(
        "interns/form.html", intern=None, stations=PIA_CITIES, form=None,
        documents_statuses=Intern.DOCUMENTS_STATUSES, certificate_statuses=Intern.CERTIFICATE_STATUSES,
        departments=Department.query.filter_by(is_active=True).order_by(db.func.lower(Department.name)).all(),
    )


@intern_bp.route("/edit/<int:intern_id>", methods=["GET", "POST"])
@login_required
@roles_required("Admin")
def edit_intern(intern_id):
    """Edit an existing intern's profile and account details."""
    intern = Intern.query.get_or_404(intern_id)
    user = intern.user

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        cnic = request.form.get("cnic", "").strip()
        station = request.form.get("station", "").strip()
        phone = request.form.get("phone", "").strip()
        email = request.form.get("email", "").strip().lower()
        qualification = request.form.get("qualification", "").strip()
        major = request.form.get("major", "").strip()
        university = request.form.get("university", "").strip()
        department_id_raw = request.form.get("department_id", "").strip()
        sub_department_id_raw = request.form.get("sub_department_id", "").strip()
        documents_status = request.form.get("documents_status", "").strip()
        certificate_status = request.form.get("certificate_status", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        start_date_raw = request.form.get("internship_start_date", "")
        end_date_raw = request.form.get("internship_end_date", "")
        photo = request.files.get("profile_picture")

        errors = []
        required_fields = {
            "Full Name": full_name,
            "CNIC": cnic,
            "Station": station,
            "Cell No": phone,
            "Email": email,
            "Qualification": qualification,
            "Major": major,
            "University": university,
            "Department": department_id_raw,
            "Username": username,
            "Internship start date": start_date_raw,
            "Internship end date": end_date_raw,
        }
        for label, value in required_fields.items():
            if not value:
                errors.append(f"{label} is required.")

        department = None
        if department_id_raw:
            try:
                department = Department.query.get(int(department_id_raw))
            except ValueError:
                department = None
            if department is None:
                errors.append("Please select a valid department from the list.")

        sub_department = None
        if sub_department_id_raw:
            try:
                sub_department = SubDepartment.query.get(int(sub_department_id_raw))
            except ValueError:
                sub_department = None
            if sub_department is None or (
                department and sub_department.department_id != department.id
            ):
                errors.append("Please select a valid sub department for the chosen department.")
                sub_department = None

        if station and station not in PIA_CITIES:
            errors.append("Please select a valid station from the list.")

        if cnic and not CNIC_RE.match(cnic.replace(" ", "")):
            errors.append("CNIC must be in the format 42101-1234567-1 (13 digits).")
        if email and not EMAIL_RE.match(email):
            errors.append("Please enter a valid email address.")
        if phone and not PHONE_RE.match(normalize_phone(phone)):
            errors.append("Please enter a valid Pakistani cell number (e.g. 03001234567).")
        if documents_status and documents_status not in Intern.DOCUMENTS_STATUSES:
            errors.append("Please select a valid Documents Status.")
        if certificate_status and certificate_status not in Intern.CERTIFICATE_STATUSES:
            errors.append("Please select a valid Certificate Status.")

        cnic_normalized = normalize_cnic(cnic) if cnic else cnic

        duplicate_email = User.query.filter(User.email == email, User.id != user.id).first()
        if duplicate_email:
            errors.append("Another user already uses this email.")
        duplicate_username = User.query.filter(
            db.func.lower(User.username) == username, User.id != user.id
        ).first()
        if duplicate_username:
            errors.append("Another user already uses this username.")
        duplicate_cnic = Intern.query.filter(
            Intern.cnic == cnic_normalized, Intern.id != intern.id
        ).first()
        if duplicate_cnic:
            errors.append("This CNIC is already registered to another intern.")

        if password and len(password) < 8:
            errors.append("New password must be at least 8 characters long.")

        start_date = end_date = None
        if start_date_raw and end_date_raw and not errors:
            try:
                start_date = _parse_date(start_date_raw)
                end_date = _parse_date(end_date_raw)
                if start_date > end_date:
                    errors.append("Internship start date must not be after the end date.")
            except ValueError:
                errors.append("Invalid date format provided.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template(
                "interns/form.html", intern=intern, stations=PIA_CITIES, form=request.form,
                documents_statuses=Intern.DOCUMENTS_STATUSES, certificate_statuses=Intern.CERTIFICATE_STATUSES,
                departments=Department.query.filter_by(is_active=True).order_by(db.func.lower(Department.name)).all(),
            )

        try:
            new_picture = save_profile_picture(photo)
            if new_picture:
                delete_profile_picture(user.profile_picture)
                user.profile_picture = new_picture

            user.email = email
            user.username = username
            if password:
                user.set_password(password)

            placement = department.name if department else None

            intern.full_name = full_name
            intern.cnic = cnic_normalized
            intern.university = university
            intern.qualification = qualification
            intern.major = major
            intern.placement = placement
            intern.department_id = department.id if department else intern.department_id
            intern.sub_department_id = sub_department.id if sub_department else None
            intern.station = station
            intern.phone = normalize_phone(phone)
            intern.documents_status = documents_status
            intern.certificate_status = certificate_status
            intern.internship_start_date = start_date
            intern.internship_end_date = end_date

            log_action(
                action="UPDATE",
                description=f"Updated intern '{full_name}' (CNIC {cnic_normalized}).",
                target_type="Intern",
                target_id=intern.id,
            )
            db.session.commit()
            flash(f"Intern '{full_name}' updated successfully.", "success")
            return redirect(url_for("intern.list_interns"))

        except ValueError as ve:
            db.session.rollback()
            flash(str(ve), "danger")
        except IntegrityError:
            db.session.rollback()
            flash("Could not update intern due to a database error.", "danger")
        except Exception:
            db.session.rollback()
            current_app.logger.exception("Failed to update intern #%s.", intern.id)
            flash("Could not update intern due to a system error. Please try again.", "danger")

    return render_template(
        "interns/form.html", intern=intern, stations=PIA_CITIES, form=None,
        documents_statuses=Intern.DOCUMENTS_STATUSES, certificate_statuses=Intern.CERTIFICATE_STATUSES,
        departments=Department.query.filter_by(is_active=True).order_by(db.func.lower(Department.name)).all(),
    )


@intern_bp.route("/toggle-status/<int:intern_id>", methods=["POST"])
@login_required
@roles_required("Admin")
def toggle_intern_status(intern_id):
    """Disable or re-enable an intern's login account (soft delete).
    The intern's record and all related history (attendance, projects,
    evaluations, etc.) remain in the database; a disabled account simply
    cannot log in, matching the Project Manager module's behaviour."""
    intern = Intern.query.get_or_404(intern_id)
    user = intern.user

    try:
        user.is_active_account = not user.is_active_account
        state = "enabled" if user.is_active_account else "disabled"
        log_action(
            action="UPDATE",
            description=f"Intern '{intern.full_name}' account {state}.",
            target_type="Intern",
            target_id=intern.id,
        )
        db.session.commit()
        try:
            from services.email_service import send_account_status_email

            send_account_status_email(user=user, is_active=user.is_active_account)
        except Exception:
            current_app.logger.exception(
                "Failed to send account status email for intern #%s.", intern.id
            )
        flash(f"Intern '{intern.full_name}' has been {state}.", "success")
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Failed to toggle status for intern #%s.", intern_id)
        flash("Could not update the intern's account status due to a system error. Please try again.", "danger")

    return redirect(url_for("intern.list_interns"))


@intern_bp.route("/bulk-status", methods=["POST"])
@login_required
@roles_required("Admin")
def bulk_toggle_intern_status():
    """Disable or enable multiple interns' login accounts in one click.
    Reuses the same soft-delete mechanism as toggle_intern_status: the
    intern's record and all related history stay in the database, only
    the linked login account's active flag is flipped."""
    intern_ids = request.form.getlist("intern_ids")
    action = (request.form.get("action") or "").strip().lower()

    if not intern_ids:
        flash("Please select at least one intern before continuing.", "warning")
        return redirect(url_for("intern.list_interns"))

    if action not in ("disable", "enable"):
        flash("Unknown bulk action requested.", "danger")
        return redirect(url_for("intern.list_interns"))

    want_active = action == "enable"
    updated = 0
    skipped = 0

    try:
        interns = Intern.query.filter(Intern.id.in_(intern_ids)).all()
        for intern in interns:
            user = intern.user
            if user is None:
                skipped += 1
                continue
            if user.is_active_account == want_active:
                continue
            user.is_active_account = want_active
            updated += 1
            try:
                from services.email_service import send_account_status_email

                send_account_status_email(user=user, is_active=user.is_active_account)
            except Exception:
                current_app.logger.exception(
                    "Failed to send account status email for intern #%s.", intern.id
                )

        log_action(
            action="UPDATE",
            description=f"Bulk {action}d {updated} intern account(s).",
            target_type="Intern",
            target_id=None,
        )
        db.session.commit()

        if updated:
            flash(f"{updated} intern(s) {action}d successfully.", "success")
        else:
            flash("No changes were made — selected interns already had that status.", "info")
        if skipped:
            flash(f"{skipped} selected intern(s) had no linked account and were skipped.", "warning")
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Failed to bulk %s interns.", action)
        flash("Could not update the selected interns due to a system error. Please try again.", "danger")

    return redirect(url_for("intern.list_interns"))


# ---------------------------------------------------------------------
# Extend Internship Duration
# ---------------------------------------------------------------------
@intern_bp.route("/extend/<int:intern_id>", methods=["GET", "POST"])
@login_required
@roles_required("Station HR", "Admin")
def extend_intern(intern_id):
    """Extend an intern's internship end date, with a required reason.
    Recorded on the intern's profile, logged to the audit trail, and
    notified to the intern (in-app + HR email) so it is fully traceable
    -- unlike silently editing the end date via the generic Edit form."""
    intern = Intern.query.get_or_404(intern_id)

    is_city_scoped, hr_city = hr_city_scope()
    if is_city_scoped and intern.station != hr_city:
        flash("You do not have permission to manage interns outside your assigned city.", "danger")
        return redirect(url_for("intern.list_interns"))

    if request.method == "POST":
        new_end_date_raw = request.form.get("new_end_date", "")
        reason = request.form.get("reason", "").strip()

        errors = []
        if not new_end_date_raw:
            errors.append("New internship end date is required.")
        if not reason:
            errors.append("A reason for the extension is required.")

        new_end_date = None
        if new_end_date_raw and not errors:
            try:
                new_end_date = _parse_date(new_end_date_raw)
            except ValueError:
                errors.append("Invalid date format provided.")

        if new_end_date and new_end_date <= intern.internship_end_date:
            errors.append("The new end date must be later than the current internship end date.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("interns/extend_form.html", intern=intern, form=request.form)

        try:
            old_end_date = intern.internship_end_date
            intern.internship_end_date = new_end_date
            # Extending un-ends/un-completes an internship that had
            # already run its course or been marked ended.
            intern.internship_status = "Active"

            log_action(
                action="EXTEND",
                description=(
                    f"Extended internship for '{intern.full_name}' from "
                    f"{old_end_date.strftime('%d %b %Y')} to {new_end_date.strftime('%d %b %Y')}. "
                    f"Reason: {reason}"
                ),
                target_type="Intern",
                target_id=intern.id,
            )
            db.session.flush()

            notify_user(
                user_id=intern.user_id,
                message=(
                    f"Your internship has been extended to {new_end_date.strftime('%d %b %Y')}."
                ),
                icon="bi-calendar-plus",
                notification_type="General",
            )
            db.session.commit()

            send_hr_pm_notification_email(
                recipients=get_hr_recipients(),
                recipient_name="HR Team",
                event_title="Internship Extended",
                event_message=f"The internship for {intern.full_name} has been extended.",
                details=[
                    ("Intern", intern.full_name),
                    ("Previous End Date", old_end_date.strftime("%d %b %Y")),
                    ("New End Date", new_end_date.strftime("%d %b %Y")),
                    ("Reason", reason),
                ],
            )

            flash(
                f"Internship for '{intern.full_name}' extended to "
                f"{new_end_date.strftime('%d %b %Y')}.",
                "success",
            )
            return redirect(url_for("intern.view_intern", intern_id=intern.id))
        except Exception:
            db.session.rollback()
            current_app.logger.exception("Failed to extend internship for intern #%s.", intern.id)
            flash("Could not extend the internship due to a system error. Please try again.", "danger")

    return render_template("interns/extend_form.html", intern=intern, form=None)


# ---------------------------------------------------------------------
# End Internship (early termination)
# ---------------------------------------------------------------------
@intern_bp.route("/end/<int:intern_id>", methods=["GET", "POST"])
@login_required
@roles_required("Station HR", "Admin")
def end_intern(intern_id):
    """End an intern's internship early. Sets the internship status to
    'Ended', caps the end date at today (never extends it), and records
    a mandatory reason -- fully audited and notified, distinct from
    deleting the intern's account/profile entirely."""
    from utils import today_pkt

    intern = Intern.query.get_or_404(intern_id)

    is_city_scoped, hr_city = hr_city_scope()
    if is_city_scoped and intern.station != hr_city:
        flash("You do not have permission to manage interns outside your assigned city.", "danger")
        return redirect(url_for("intern.list_interns"))

    if request.method == "POST":
        reason = request.form.get("reason", "").strip()
        errors = []
        if not reason:
            errors.append("A reason for ending the internship is required.")
        if intern.effective_status == "Ended":
            errors.append("This internship has already been ended.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("interns/end_form.html", intern=intern, form=request.form)

        try:
            today = today_pkt()
            if intern.internship_end_date > today:
                intern.internship_end_date = today
            intern.internship_status = "Ended"
            intern.end_reason = reason

            log_action(
                action="END",
                description=f"Ended internship for '{intern.full_name}' early. Reason: {reason}",
                target_type="Intern",
                target_id=intern.id,
            )
            db.session.flush()

            notify_user(
                user_id=intern.user_id,
                message="Your internship has been ended by HR. Contact HR for details.",
                icon="bi-flag",
                notification_type="General",
            )
            db.session.commit()

            send_hr_pm_notification_email(
                recipients=get_hr_recipients(),
                recipient_name="HR Team",
                event_title="Internship Ended Early",
                event_message=f"The internship for {intern.full_name} has been ended early by HR.",
                details=[
                    ("Intern", intern.full_name),
                    ("Effective End Date", intern.internship_end_date.strftime("%d %b %Y")),
                    ("Reason", reason),
                ],
            )

            flash(f"Internship for '{intern.full_name}' has been ended.", "success")
            return redirect(url_for("intern.view_intern", intern_id=intern.id))
        except Exception:
            db.session.rollback()
            current_app.logger.exception("Failed to end internship for intern #%s.", intern.id)
            flash("Could not end the internship due to a system error. Please try again.", "danger")

    return render_template("interns/end_form.html", intern=intern, form=None)


# ---------------------------------------------------------------------
# Download Intern Profile (Document)
# ---------------------------------------------------------------------
@intern_bp.route("/profile-pdf/<int:intern_id>")
@login_required
@roles_required("Station HR", "Admin")
def download_profile_pdf(intern_id):
    """Download a single intern's profile as a branded PDF document."""
    from services import pdf_reports

    intern = Intern.query.get_or_404(intern_id)
    is_city_scoped, hr_city = hr_city_scope()
    if is_city_scoped and intern.station != hr_city:
        flash("You do not have permission to access interns outside your assigned city.", "danger")
        return redirect(url_for("intern.list_interns"))
    try:
        buffer = pdf_reports.build_intern_profile_pdf(intern)
        return send_file(
            buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"intern_profile_{intern.full_name.replace(' ', '_')}.pdf",
        )
    except Exception:
        current_app.logger.exception("Failed to generate profile PDF for intern #%s.", intern_id)
        flash("Could not generate the profile document due to a system error. Please try again.", "danger")
        return redirect(url_for("intern.view_intern", intern_id=intern_id))


# ======================================================================
# Bulk Import Interns (Admin only) -- Excel (.xlsx) / CSV upload
# ======================================================================

# Exact column order from the FTOP Excel import template.
# NOTE: the old free-text "Placement" column was replaced with two
# structured columns -- "Department" and "Division/Section" -- that
# must match an existing Department -> SubDepartment pair exactly
# (case insensitive), the same hierarchy used by every cascading
# dropdown elsewhere in the app (see utils.DEPARTMENT_HIERARCHY / the
# department_select component). Free-text Placement values are no
# longer accepted or auto-created into ad-hoc Department rows.
IMPORT_COLUMNS = [
    "Full Name",
    "CNIC",
    "Station",
    "Cell no",
    "Email",
    "Qualification",
    "Major",
    "University",
    "Department",
    "Division/Section",
    "Internship Start Date",
    "Internship End Date",
    "Documents Status",
    "Certificate Status",
]

# Older uploaded templates may still use the previous column name --
# accepted transparently so existing import files keep working.
_LEGACY_COLUMN_ALIASES = {
    "Sub Department": "Division/Section",
}

_IMPORT_DATE_FORMATS = [
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%Y/%m/%d",
    "%d %b %Y",
    "%d %B %Y",
]


class _RowError(Exception):
    """Raised for a single row that fails validation/duplicate checks
    during bulk import. Caught by the row loop so the remaining rows
    keep processing. `category` is one of "duplicate_cnic",
    "duplicate_email", or "validation", used to build the import
    summary's breakdown counts."""

    def __init__(self, message, category="validation"):
        super().__init__(message)
        self.category = category


def _to_text(value) -> str:
    """Normalise a raw cell value (str, number, None, ...) to a
    trimmed string for text fields."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_import_date(value):
    """Parse a date cell that may already be a date/datetime object
    (typical for .xlsx) or a plain string (typical for .csv / text
    cells), trying a handful of common formats. Returns None if the
    value is missing or unparsable."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in _IMPORT_DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _read_xlsx_rows(file_storage):
    """Read an uploaded .xlsx file into a list of dict rows keyed by
    the expected import columns. Returns (rows, error_message)."""
    try:
        wb = load_workbook(file_storage, data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
    except Exception:
        return None, "The uploaded file is not a valid .xlsx file."

    if header is None:
        return None, "The uploaded file is empty."

    header = [_to_text(h) for h in header]
    header = [_LEGACY_COLUMN_ALIASES.get(h, h) for h in header]
    missing = [c for c in IMPORT_COLUMNS if c not in header]
    if missing:
        return None, f"Missing required column(s): {', '.join(missing)}."

    rows = []
    for values in rows_iter:
        if values is None or all(v is None or _to_text(v) == "" for v in values):
            continue
        row = {}
        for col_name, val in zip(header, values):
            if col_name in IMPORT_COLUMNS:
                row[col_name] = val
        rows.append(row)
    return rows, None


def _read_csv_rows(file_storage):
    """Read an uploaded .csv file into a list of dict rows keyed by
    the expected import columns. Returns (rows, error_message)."""
    raw = file_storage.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = raw.decode("latin-1")
        except Exception:
            return None, "Could not read the uploaded CSV file. Please save it as UTF-8 CSV."

    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return None, "The uploaded file is empty."

    header = [_to_text(h) for h in header]
    header = [_LEGACY_COLUMN_ALIASES.get(h, h) for h in header]
    missing = [c for c in IMPORT_COLUMNS if c not in header]
    if missing:
        return None, f"Missing required column(s): {', '.join(missing)}."

    rows = []
    for values in reader:
        if not values or all(_to_text(v) == "" for v in values):
            continue
        row = {}
        for col_name, val in zip(header, values):
            if col_name in IMPORT_COLUMNS:
                row[col_name] = val
        rows.append(row)
    return rows, None


def _import_single_row(row, seen_emails, seen_usernames, seen_cnics):
    """Validate one Excel/CSV row and create its User + Intern records
    via the same _create_intern_account() helper used by the manual
    Add Intern form. Raises _RowError with a human-readable reason on
    any validation/duplicate failure; the caller rolls back just this
    row and continues with the rest of the file.

    Username = Email, Password = CNIC for every imported intern (per
    spec); no Username/Password columns exist in the template and no
    welcome email is sent for bulk imports."""
    full_name = _to_text(row.get("Full Name"))
    cnic_raw = _to_text(row.get("CNIC"))
    station = _to_text(row.get("Station"))
    phone_raw = _to_text(row.get("Cell no"))
    email = _to_text(row.get("Email")).lower()
    qualification = _to_text(row.get("Qualification"))
    major = _to_text(row.get("Major"))
    university = _to_text(row.get("University"))
    department_name = _to_text(row.get("Department"))
    sub_department_name = _to_text(row.get("Division/Section"))
    documents_status = _to_text(row.get("Documents Status")) or "Pending"
    certificate_status = _to_text(row.get("Certificate Status")) or "Pending"

    # ---- Required fields ----
    required = {
        "Full Name": full_name,
        "CNIC": cnic_raw,
        "Station": station,
        "Cell no": phone_raw,
        "Email": email,
        "Qualification": qualification,
        "Major": major,
        "University": university,
        "Department": department_name,
        "Division/Section": sub_department_name,
    }
    missing = [label for label, value in required.items() if not value]
    if missing:
        raise _RowError(f"Missing required field(s): {', '.join(missing)}.")

    # ---- Department / Division/Section (structured, no free-text) ----
    department = Department.query.filter(
        db.func.lower(Department.name) == department_name.lower(),
        Department.is_active.is_(True),
    ).first()
    if department is None:
        raise _RowError(
            f"Unknown Department '{department_name}'. It must match an existing, active department exactly."
        )

    sub_department = SubDepartment.query.filter(
        db.func.lower(SubDepartment.name) == sub_department_name.lower(),
        SubDepartment.department_id == department.id,
        SubDepartment.is_active.is_(True),
    ).first()
    if sub_department is None:
        raise _RowError(
            f"Unknown Division/Section '{sub_department_name}' under Department '{department_name}'. "
            "It must match an existing, active division/section exactly."
        )

    # ---- Format validation ----
    if not EMAIL_RE.match(email):
        raise _RowError(f"Invalid Email format: '{email}'.")
    if not CNIC_RE.match(cnic_raw.replace(" ", "")):
        raise _RowError(f"Invalid CNIC format: '{cnic_raw}'. Expected 13 digits (e.g. 42101-1234567-1).")
    phone = normalize_phone(phone_raw)
    if not PHONE_RE.match(phone):
        raise _RowError(f"Invalid Cell no format: '{phone_raw}'. Expected a Pakistani mobile number.")
    if station not in PIA_CITIES:
        raise _RowError(f"Unknown Station '{station}'. Must be one of: {', '.join(PIA_CITIES)}.")
    if documents_status not in Intern.DOCUMENTS_STATUSES:
        raise _RowError(
            f"Invalid Documents Status '{documents_status}'. Must be one of: {', '.join(Intern.DOCUMENTS_STATUSES)}."
        )
    if certificate_status not in Intern.CERTIFICATE_STATUSES:
        raise _RowError(
            f"Invalid Certificate Status '{certificate_status}'. Must be one of: {', '.join(Intern.CERTIFICATE_STATUSES)}."
        )

    cnic = normalize_cnic(cnic_raw)

    start_date = _parse_import_date(row.get("Internship Start Date"))
    end_date = _parse_import_date(row.get("Internship End Date"))
    if not start_date:
        raise _RowError("Invalid or missing Internship Start Date.")
    if not end_date:
        raise _RowError("Invalid or missing Internship End Date.")
    if start_date > end_date:
        raise _RowError("Internship Start Date must not be after the End Date.")

    # Username = Email, Password = CNIC (auto-generated, never taken
    # from the file -- the spec forbids Username/Password columns).
    username = email
    password = cnic

    if email in seen_emails or User.query.filter_by(email=email).first():
        raise _RowError(f"Duplicate Email: '{email}' already exists.", category="duplicate_email")
    if username in seen_usernames or User.query.filter_by(username=username).first():
        raise _RowError(f"Duplicate Email/Username: '{username}' already exists.", category="duplicate_email")
    if cnic in seen_cnics or Intern.query.filter_by(cnic=cnic).first():
        raise _RowError(f"Duplicate CNIC: '{cnic}' already exists.", category="duplicate_cnic")

    _create_intern_account(
        full_name=full_name,
        cnic=cnic,
        university=university,
        qualification=qualification,
        major=major,
        placement=None,
        station=station,
        phone=phone,
        email=email,
        username=username,
        password=password,
        start_date=start_date,
        end_date=end_date,
        documents_status=documents_status,
        certificate_status=certificate_status,
        department=department,
        sub_department_id=sub_department.id,
        # Interns imported in bulk log in with their CNIC as password,
        # so force them to set their own on first login. No email is
        # sent for bulk imports (see import_interns() below).
        force_password_reset=True,
    )

    # Only mark as "seen" once the row has actually succeeded, so a
    # failed row doesn't block a later, legitimately different row.
    seen_emails.add(email)
    seen_usernames.add(username)
    seen_cnics.add(cnic)


@intern_bp.route("/import", methods=["GET", "POST"])
@login_required
@roles_required("Admin")
def import_interns():
    """Bulk-import interns from an uploaded .xlsx/.csv file. Every row
    is processed independently (its own DB savepoint) so a single bad
    row is skipped-and-reported instead of failing the whole file."""
    if request.method == "GET":
        return render_template("interns/import.html", result=None)

    upload = request.files.get("import_file")
    if not upload or upload.filename == "":
        flash("Please choose an Excel (.xlsx) or CSV file to upload.", "danger")
        return redirect(url_for("intern.import_interns"))

    filename = upload.filename.lower()
    if filename.endswith(".xlsx"):
        rows, header_error = _read_xlsx_rows(upload)
    elif filename.endswith(".csv"):
        rows, header_error = _read_csv_rows(upload)
    else:
        flash("Unsupported file type. Please upload a .xlsx or .csv file.", "danger")
        return redirect(url_for("intern.import_interns"))

    if header_error:
        flash(header_error, "danger")
        return redirect(url_for("intern.import_interns"))

    if not rows:
        flash("The uploaded file doesn't contain any data rows.", "warning")
        return redirect(url_for("intern.import_interns"))

    seen_emails, seen_usernames, seen_cnics = set(), set(), set()
    errors = []  # list of (row_number, message, category)
    duplicate_cnic_count = 0
    duplicate_email_count = 0
    validation_error_count = 0
    success_count = 0

    for idx, row in enumerate(rows, start=2):  # row 1 in the file is the header
        try:
            with db.session.begin_nested():
                _import_single_row(row, seen_emails, seen_usernames, seen_cnics)
            success_count += 1
        except _RowError as exc:
            errors.append((idx, str(exc), exc.category))
            if exc.category == "duplicate_cnic":
                duplicate_cnic_count += 1
            elif exc.category == "duplicate_email":
                duplicate_email_count += 1
            else:
                validation_error_count += 1
        except IntegrityError:
            errors.append((idx, "This row conflicts with an existing record (duplicate CNIC/email/username).", "duplicate_cnic"))
            duplicate_cnic_count += 1
        except Exception:
            current_app.logger.exception("Bulk intern import: unexpected error on row %s.", idx)
            errors.append((idx, "Unexpected error while processing this row.", "validation"))
            validation_error_count += 1

    if success_count:
        log_action(
            action="IMPORT",
            description=f"Admin imported {success_count} intern(s) via Excel.",
            target_type="Intern",
        )

    db.session.commit()

    result = {
        "total": len(rows),
        "success": success_count,
        "skipped": len(errors),
        "duplicate_cnic": duplicate_cnic_count,
        "duplicate_email": duplicate_email_count,
        "validation_errors": validation_error_count,
        "errors": errors,
    }

    # Stashed for the "Download Error Report" button -- small enough
    # (row number + short message) to keep in the session safely.
    session["last_import_errors"] = [{"row": r, "message": m} for r, m, _c in errors]

    if success_count:
        flash(f"Import complete: {success_count} of {len(rows)} intern(s) imported successfully.", "success")
    else:
        flash("Import complete: no rows could be imported. See the error report below.", "warning")

    return render_template("interns/import.html", result=result)


@intern_bp.route("/import/error-report")
@login_required
@roles_required("Admin")
def import_error_report():
    """Download the failed rows + reasons from the most recent bulk
    import as an .xlsx error report."""
    errors = session.get("last_import_errors") or []
    if not errors:
        flash("No recent import errors to download.", "info")
        return redirect(url_for("intern.import_interns"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Errors"
    ws.append(["Row", "Reason"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for item in errors:
        ws.append([item["row"], item["message"]])
    ws.column_dimensions[get_column_letter(1)].width = 10
    ws.column_dimensions[get_column_letter(2)].width = 80

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="intern_import_errors.xlsx",
    )


@intern_bp.route("/import/template")
@login_required
@roles_required("Admin")
def import_template():
    """Download a blank .xlsx template with exactly the columns the
    bulk importer expects (no Username/Password columns)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Interns"
    ws.append(IMPORT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    column_widths = [22, 18, 20, 16, 26, 18, 18, 22, 22, 22, 20, 20, 18, 18]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="intern_import_template.xlsx",
    )


@intern_bp.route("/export")
@login_required
@roles_required("Station HR", "Admin")
def export_interns():
    """Export every intern to .xlsx using the exact same column order
    as the Excel import template, so the file can be re-imported
    elsewhere (or into another environment) unchanged."""
    interns = Intern.query
    is_city_scoped, hr_city = hr_city_scope()
    if is_city_scoped:
        interns = interns.filter(Intern.station == hr_city)
    interns = interns.order_by(Intern.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Interns"
    ws.append(IMPORT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for intern in interns:
        ws.append([
            intern.full_name,
            intern.cnic,
            intern.station,
            intern.phone,
            intern.user.email if intern.user else "",
            intern.qualification,
            intern.major or "",
            intern.university,
            intern.department.name if intern.department else "",
            intern.sub_department.name if intern.sub_department else "",
            intern.internship_start_date.strftime("%Y-%m-%d") if intern.internship_start_date else "",
            intern.internship_end_date.strftime("%Y-%m-%d") if intern.internship_end_date else "",
            intern.documents_status,
            intern.certificate_status,
        ])

    column_widths = [22, 18, 20, 16, 26, 18, 18, 22, 22, 22, 20, 20, 18, 18]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_action(
        action="EXPORT",
        description=f"Exported {len(interns)} intern(s) to Excel.",
        target_type="Intern",
    )
    db.session.commit()

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="interns_export.xlsx",
    )
