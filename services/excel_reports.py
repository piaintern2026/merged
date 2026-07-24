"""
services/excel_reports.py
---------------------------
Generates the same five reports as services/pdf_reports.py, but as
.xlsx workbooks using OpenPyXL. Every function returns an in-memory
BytesIO buffer ready to be sent with Flask's send_file().
"""

import os
from datetime import datetime
from utils import now_pkt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
LOGO_PATH = os.path.join(BASE_DIR, "static", "images", "pia-logo.png")

REPORT_ORG_NAME = "Pakistan International Airlines"
REPORT_SYSTEM_NAME = "Intern Management System"
REPORT_CONFIDENTIALITY_NOTE = "Confidential - For Internal Management Use Only"

# PIA brand colors (matching static/css/style.css), as ARGB hex for OpenPyXL
BANNER_FILL = PatternFill(start_color="FF012A54", end_color="FF012A54", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFC9A24B", end_color="FFC9A24B", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FF012A54", end_color="FF012A54", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True, size=11)
STRIPE_FILL = PatternFill(start_color="FFEEF1F5", end_color="FFEEF1F5", fill_type="solid")
TITLE_FONT = Font(color="FF012A54", bold=True, size=16)
SUBTITLE_FONT = Font(color="FF6C757D", italic=True, size=10)
BANNER_ORG_FONT = Font(color="FFFFFFFF", bold=True, size=14)
BANNER_SUB_FONT = Font(color="FFCFE0EE", size=9)
FOOTER_FONT = Font(color="FF6C757D", italic=True, size=8)
THIN_BORDER = Border(
    left=Side(style="thin", color="FFDFE4EA"),
    right=Side(style="thin", color="FFDFE4EA"),
    top=Side(style="thin", color="FFDFE4EA"),
    bottom=Side(style="thin", color="FFDFE4EA"),
)


def _new_sheet(title: str, subtitle: str, header: list):
    """Create a workbook + worksheet with a PIA-styled letterhead
    (logo + navy banner), title block, and header row already
    written. Returns (workbook, worksheet, next_row)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    num_cols = max(len(header), 4)
    last_col_letter = get_column_letter(num_cols)

    # ---- Letterhead banner (rows 1-3): navy fill across all columns ----
    for row in (1, 2, 3):
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = BANNER_FILL
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws.merge_cells(f"A3:{last_col_letter}3")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 16

    ws["A1"] = f"      {REPORT_ORG_NAME}"
    ws["A1"].font = BANNER_ORG_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"] = f"      {REPORT_SYSTEM_NAME}"
    ws["A2"].font = BANNER_SUB_FONT
    ws["A2"].alignment = Alignment(vertical="center")

    # Gold accent rule under the banner
    ws.row_dimensions[3].height = 4
    for col in range(1, num_cols + 1):
        ws.cell(row=3, column=col).fill = GOLD_FILL

    if os.path.exists(LOGO_PATH):
        try:
            img = XLImage(LOGO_PATH)
            img.height = 42
            img.width = 42
            ws.add_image(img, "A1")
        except Exception:
            pass

    # ---- Title block ----
    ws["A5"] = title
    ws["A5"].font = TITLE_FONT
    ws["A6"] = f"{subtitle} — Generated {now_pkt().strftime('%d %b %Y, %I:%M %p')}"
    ws["A6"].font = SUBTITLE_FONT

    header_row = 8
    for col_index, col_name in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    return wb, ws, header_row + 1


def _write_footer(ws, start_row: int, num_cols: int) -> None:
    """Write a confidentiality footer note a couple of rows below the
    last data row, matching the PDF report's footer branding."""
    footer_row = start_row + 2
    last_col_letter = get_column_letter(max(num_cols, 1))
    ws.merge_cells(f"A{footer_row}:{last_col_letter}{footer_row}")
    ws[f"A{footer_row}"] = f"{REPORT_CONFIDENTIALITY_NOTE}  |  {REPORT_ORG_NAME}"
    ws[f"A{footer_row}"].font = FOOTER_FONT


def _write_rows(ws, rows: list, start_row: int, num_cols: int) -> None:
    """Write data rows below the header, with alternating row shading
    and auto-sized columns."""
    for row_offset, row_data in enumerate(rows):
        row_index = start_row + row_offset
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_offset % 2 == 1:
                cell.fill = STRIPE_FILL

    # Auto-size columns based on content length (capped for readability).
    for col_index in range(1, num_cols + 1):
        max_len = 10
        for row in ws.iter_rows(min_col=col_index, max_col=col_index):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_index)].width = min(max_len + 2, 45)


def _finalize(wb: Workbook) -> BytesIO:
    """Save the workbook into an in-memory buffer, rewound to 0."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ------------------------------------------------------------------------
# 1. Attendance Report
# ------------------------------------------------------------------------
def build_attendance_excel(records: list) -> BytesIO:
    header = ["Intern", "Department", "Date", "Time", "Status", "Marked By", "Remarks"]
    wb, ws, start_row = _new_sheet("Attendance Report", f"{len(records)} record(s)", header)

    rows = [
        [
            r.intern.full_name,
            r.intern.department.name,
            r.date.strftime("%d %b %Y"),
            r.time.strftime("%I:%M %p") if r.time else "-",
            r.status,
            r.marked_by.full_name,
            r.remarks or "-",
        ]
        for r in records
    ]
    _write_rows(ws, rows, start_row, len(header))
    _write_footer(ws, start_row + len(rows), len(header))
    return _finalize(wb)


# ------------------------------------------------------------------------
# 2. Evaluation Report
# ------------------------------------------------------------------------
def build_evaluation_excel(evaluations: list) -> BytesIO:
    header = [
        "Intern", "Type", "Technical", "Communication", "Discipline",
        "Learning", "Teamwork", "Attendance", "Total /60", "Percentage", "Evaluated By", "Remarks",
    ]
    wb, ws, start_row = _new_sheet("Evaluation Report", f"{len(evaluations)} evaluation(s)", header)

    rows = [
        [
            e.intern.full_name,
            e.evaluation_type,
            e.technical_skills,
            e.communication,
            e.discipline,
            e.learning,
            e.teamwork,
            e.attendance_score,
            e.total_score,
            f"{e.percentage}%",
            e.evaluated_by.display_name(),
            e.remarks or "-",
        ]
        for e in evaluations
    ]
    _write_rows(ws, rows, start_row, len(header))
    _write_footer(ws, start_row + len(rows), len(header))
    return _finalize(wb)


# ------------------------------------------------------------------------
# 3. Intern Progress Report (single intern, multi-section)
# ------------------------------------------------------------------------
def build_intern_progress_excel(
    intern, attendance_percentage: float, work_logs: list, evaluations: list, submissions: list
) -> BytesIO:
    header = ["Field", "Value"]
    wb, ws, start_row = _new_sheet(
        "Intern Progress Report", intern.full_name, header
    )

    profile_rows = [
        ["Department", intern.department.name],
        ["CNIC", intern.cnic],
        ["University", intern.university],
        ["Degree", intern.degree],
        ["Internship Start", intern.internship_start_date.strftime("%d %b %Y")],
        ["Internship End", intern.internship_end_date.strftime("%d %b %Y")],
        ["Attendance %", f"{attendance_percentage}%"],
    ]
    _write_rows(ws, profile_rows, start_row, len(header))

    # Work Log sheet
    ws_logs = wb.create_sheet("Work Logs")
    log_header = ["Date", "Description", "Hours", "Progress %"]
    for col_index, col_name in enumerate(log_header, start=1):
        cell = ws_logs.cell(row=1, column=col_index, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    log_rows = [
        [wl.log_date.strftime("%d %b %Y"), wl.description, wl.hours_worked, wl.progress_percent]
        for wl in work_logs
    ]
    _write_rows(ws_logs, log_rows, 2, len(log_header))

    # Evaluations sheet
    ws_eval = wb.create_sheet("Evaluations")
    eval_header = ["Type", "Date", "Total /60", "Percentage", "Remarks"]
    for col_index, col_name in enumerate(eval_header, start=1):
        cell = ws_eval.cell(row=1, column=col_index, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    eval_rows = [
        [e.evaluation_type, e.created_at.strftime("%d %b %Y"), e.total_score, f"{e.percentage}%", e.remarks or "-"]
        for e in evaluations
    ]
    _write_rows(ws_eval, eval_rows, 2, len(eval_header))

    # Submissions sheet
    ws_sub = wb.create_sheet("Submissions")
    sub_header = ["Link", "Submitted"]
    for col_index, col_name in enumerate(sub_header, start=1):
        cell = ws_sub.cell(row=1, column=col_index, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    sub_rows = [
        [s.link, s.submitted_at.strftime("%d %b %Y")]
        for s in submissions
    ]
    _write_rows(ws_sub, sub_rows, 2, len(sub_header))
    _write_footer(ws, start_row + len(profile_rows), len(header))

    return _finalize(wb)


# ------------------------------------------------------------------------
# 4. Department Summary Report
# ------------------------------------------------------------------------
def build_department_summary_excel(department_rows: list) -> BytesIO:
    header = ["Department", "Status", "Project Managers", "Interns", "Projects", "Avg. Evaluation %"]
    wb, ws, start_row = _new_sheet(
        "Department Summary Report", f"{len(department_rows)} department(s)", header
    )

    rows = [
        [
            d["name"],
            d["status"],
            d["pm_count"],
            d["intern_count"],
            d["project_count"],
            f"{d['avg_score']}%" if d["avg_score"] is not None else "-",
        ]
        for d in department_rows
    ]
    _write_rows(ws, rows, start_row, len(header))
    _write_footer(ws, start_row + len(rows), len(header))
    return _finalize(wb)


# ------------------------------------------------------------------------
# 6. Station x Department Report (City x Department Matrix)
# ------------------------------------------------------------------------
def build_station_department_excel(matrix_data: dict) -> BytesIO:
    cities = matrix_data["cities"]
    matrix = matrix_data["matrix"]
    header = ["Department"] + cities + ["Total"]
    wb, ws, start_row = _new_sheet(
        "Station \u00d7 Department Report", f"{matrix_data['grand_total']} intern(s) total", header
    )

    rows = [[row["department"]] + row["cells"] + [row["row_total"]] for row in matrix]
    _write_rows(ws, rows, start_row, len(header))

    # Totals row, styled like the header row's tfoot on screen.
    total_row_index = start_row + len(rows)
    total_row_data = ["Total"] + matrix_data["city_totals"] + [matrix_data["grand_total"]]
    for col_index, value in enumerate(total_row_data, start=1):
        cell = ws.cell(row=total_row_index, column=col_index, value=value)
        cell.border = THIN_BORDER
        cell.font = Font(bold=True)
        cell.fill = STRIPE_FILL

    _write_footer(ws, total_row_index + 1, len(header))
    return _finalize(wb)


# ------------------------------------------------------------------------
# 5. Project Summary Report
# ------------------------------------------------------------------------
def build_project_summary_excel(projects: list) -> BytesIO:
    header = ["Title", "Department", "Manager", "Intern", "Priority", "Status", "Start Date", "Deadline"]
    wb, ws, start_row = _new_sheet(
        "Project Summary Report", f"{len(projects)} project(s)", header
    )

    rows = [
        [
            p.title,
            p.department.name,
            p.manager.full_name if p.manager else "-",
            p.intern.full_name if p.intern else "-",
            p.priority,
            p.status,
            p.start_date.strftime("%d %b %Y"),
            p.deadline.strftime("%d %b %Y"),
        ]
        for p in projects
    ]
    _write_rows(ws, rows, start_row, len(header))
    _write_footer(ws, start_row + len(rows), len(header))
    return _finalize(wb)
